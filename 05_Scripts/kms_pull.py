#!/usr/bin/env python3
"""
KMS Sheet Puller — OpenDurian How-to
ดึงข้อมูล Content จาก Google Sheet KMS แล้วอัปเดต Dashboard

Flow:
  1. Download KMS Sheet (Apr26 tab) เป็น CSV
  2. Parse content rows → ได้ video_id, creator, type, product, title
  3. ถ้ามี Supermetrics Sheet → join metrics by video_id
  4. อัปเดต TikTok section ใน dashboard HTML (RAW_DATA)
  5. อัปเดต KMS_CAL_DATA ใน dashboard HTML (Calendar view)
"""

import csv
import json
import re
import os
import io
import requests
from datetime import datetime
from collections import defaultdict

# ============================================================
# CONFIG — แก้ตรงนี้
# ============================================================
KMS_SHEET_ID  = "1HoDVCRxrSdaJOqTwaTjR3VBqKBl1hbx7CDyt8skrGL4"
KMS_TAB_GIDS  = {
    "Feb26": "",          # ใส่ GID ของ Tab Feb26 (ดูจาก URL ของ tab นั้น)
    "Mar26": "",          # ใส่ GID ของ Tab Mar26
    "Apr26": "1506419058",
}

# Supermetrics Sheet — ใส่เมื่อ Dev ตั้งให้เสร็จแล้ว
SUPERMETRICS_CSV_URL = ""   # ← เว้นว่างไว้ก่อน รอ Dev

# path dashboard
DASHBOARD_PATH = "/Users/opendurian/Documents/Claude/Projects/Excel: Content KPI Dashboard/index.html"
# ============================================================

SHEET_CSV_URL = "https://docs.google.com/spreadsheets/d/{sid}/export?format=csv&gid={gid}"
# โหลดทั้ง workbook (.xlsx) → ได้ทุกแท็บ คอลัมน์ตรง grid เป๊ะ (ไม่เหมือน gviz ที่ตัดคอลัมน์ว่าง)
SHEET_XLSX_URL = "https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"

# auto-discover: แท็บชื่อรูปแบบเดือน รองรับทั้ง "Feb26" และ "Feb 26" (มีเว้นวรรค)
TAB_MONTH_RE = re.compile(r'^[A-Za-z]{3}\s*\d{2}$')

def discover_tabs_xlsx():
    """โหลด workbook ทั้งไฟล์ → คืน {tab_label: rows(list[list[str]])}
       เฉพาะแท็บที่ชื่อเป็นเดือน (MonYY) → ครอบคลุมชีตเก่า+ใหม่อัตโนมัติ ไม่ต้องใส่ GID"""
    try:
        import openpyxl
    except ImportError:
        print("  ⚠️  ไม่มี openpyxl — ติดตั้ง: pip3 install openpyxl --break-system-packages")
        return {}
    url = SHEET_XLSX_URL.format(sid=KMS_SHEET_ID)
    try:
        r = requests.get(url, timeout=60); r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception as e:
        print(f"  ⚠️  โหลด workbook ไม่ได้: {e}")
        return {}
    out = {}
    for name in wb.sheetnames:
        nm = (name or "").strip()
        if not TAB_MONTH_RE.match(nm):
            continue
        label = nm.replace(" ", "")        # "Feb 26" → "Feb26" (canonical ให้ downstream)
        ws = wb[nm]; rows = []
        for row in ws.iter_rows(values_only=True):
            cells = []
            for c in row:
                if c is None:
                    cells.append("")
                elif hasattr(c, "strftime"):          # datetime → ISO (parse_date รองรับ)
                    cells.append(c.strftime("%Y-%m-%d"))
                else:
                    cells.append(str(c))
            rows.append(cells)
        if rows:
            out[label] = rows
    return out

CREATOR_NAMES = {"ริว", "มิ้น", "หมิว"}

# mapping: tab label → calendar key / month info
def _cal_display_key(label):
    """'Feb26' → 'Feb 26' — dynamic ไม่ hardcode ปี"""
    import re as _re
    m = _re.match(r'([A-Za-z]{3})(\d{2})', label)
    if m:
        return f"{m.group(1)} {m.group(2)}"
    return label

class _CalKeyMap:
    def get(self, key, default=None):
        result = _cal_display_key(key)
        return result if result != key else (default if default is not None else key)

CAL_KEY_MAP = _CalKeyMap()
def _cal_month_tuple(label):
    """'Feb26' → (2026, 1) — 0-indexed month สำหรับ JS Date, dynamic ไม่ hardcode"""
    import re as _re
    mon_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
               "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
    m = _re.match(r'([A-Za-z]{3})(\d{2})', label)
    if m:
        mon = mon_map.get(m.group(1), 0)
        yr = 2000 + int(m.group(2))
        if mon > 0:
            return yr, mon - 1  # 0-indexed for JS Date
    return None

class _CalMonthMap:
    def get(self, key, default=None):
        result = _cal_month_tuple(key)
        return result if result is not None else default

# (year, month 0-indexed for JS Date) — auto-generated from label
CAL_MONTH_MAP = _CalMonthMap()
MONTH_LABELS = {
    1: "มกราคม", 2: "กุมภาพันธ์", 3: "มีนาคม", 4: "เมษายน",
    5: "พฤษภาคม", 6: "มิถุนายน", 7: "กรกฎาคม", 8: "สิงหาคม",
    9: "กันยายน", 10: "ตุลาคม", 11: "พฤศจิกายน", 12: "ธันวาคม",
}

# คอลัมน์ใน KMS Sheet (0-indexed)
COL = {
    "date":         0,
    "creator":      1,
    "format":       2,   # VDO / IMG
    "product":      4,
    "content_type": 5,   # Sale / Engage / Live
    "key_message":  6,
    "published":    11,
    "pub_date":     12,
    "roas_7d":      14,
    "fb_post_url":  15,
    "ads_fb_name":  16,
    "fb_post_id":   17,
    "tiktok_url":   19,
    "tiktok_vid":   20,
}


def download_csv(url):
    try:
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        text = r.content.decode("utf-8-sig")
        return list(csv.reader(io.StringIO(text)))
    except Exception as e:
        print(f"  ⚠️  ดาวน์โหลดไม่ได้: {e}")
        return []


def clean(s):
    return s.strip().strip('"')


def parse_date(raw, year=None):
    if year is None:
        from datetime import datetime as _dt
        year = _dt.now().year
    """แปลง 'Wed,  1, Apr' → '2026-04-01' (รองรับ ISO จาก xlsx ด้วย)"""
    raw = re.sub(r'\s+', ' ', raw.strip().strip('"'))
    _iso = re.match(r'(\d{4})-(\d{2})-(\d{2})', raw)
    if _iso:
        return f"{_iso.group(1)}-{_iso.group(2)}-{_iso.group(3)}"
    parts = raw.split(",")
    if len(parts) >= 3:
        day_s  = parts[1].strip()
        mon_s  = parts[2].strip()
    elif len(parts) == 2:
        day_s  = parts[0].strip()
        mon_s  = parts[1].strip()
    else:
        return ""
    mon_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
               "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
    try:
        day = int(re.sub(r'\D','',day_s))
        mon = mon_map.get(mon_s[:3], 0)
        if mon == 0:
            return ""
        return f"{year}-{mon:02d}-{day:02d}"
    except:
        return ""


def month_key(date_str):
    """'2026-04-15' → 'Apr26' — dynamic ไม่ hardcode ปี/เดือน"""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%b") + dt.strftime("%y")
    except:
        return "Unknown"


def safe_get(row, col_idx, default=""):
    """ดึง cell ปลอดภัย ไม่ IndexError"""
    try:
        return clean(row[col_idx])
    except:
        return default


# ─── RAW_DATA (TikTok performance) ──────────────────────────

def parse_kms_tab(rows, tab_label):
    """
    อ่าน rows จาก KMS tab → list of content dicts
    (กรองเฉพาะแถวที่มี TikTok VDO ID)
    """
    contents = []

    tiktok_col = 20  # default
    if rows:
        for j, cell in enumerate(rows[0]):
            c = cell.strip().lower()
            if "tiktok" in c and "vdo" in c:
                tiktok_col = j
                break
        print(f"  TikTok VDO col = {tiktok_col} ('{rows[0][tiktok_col] if len(rows[0]) > tiktok_col else '?'}')")

    for row in rows[1:]:
        if len(row) <= tiktok_col:
            continue

        vid_id = safe_get(row, tiktok_col)
        if vid_id and not re.match(r'^\d+$', vid_id):
            vid_id = ""
        if not vid_id or vid_id in ("#N/A", "-", ""):
            continue

        creator = safe_get(row, COL["creator"])
        if creator not in CREATOR_NAMES:
            continue

        pub_status = safe_get(row, COL["published"])
        if pub_status not in ("Published", "รอ Publish"):
            continue

        date_raw = safe_get(row, COL["date"])
        pub_date = safe_get(row, COL["pub_date"])
        date_str = parse_date(pub_date) or parse_date(date_raw)

        contents.append({
            "video_id":     vid_id,
            "creator":      creator,
            "content_format": safe_get(row, COL["format"]),
            "product":      safe_get(row, COL["product"]),
            "content_type": safe_get(row, COL["content_type"]),
            "content_name": safe_get(row, COL["key_message"]),
            "date":         date_str,
            "month":        month_key(date_str) if date_str else tab_label,
            "tiktok_url":   safe_get(row, COL["tiktok_url"]),
            "fb_post_id":   safe_get(row, COL["fb_post_id"]),
            "ads_fb_name":  safe_get(row, COL["ads_fb_name"]),
            "roas_7d":      safe_get(row, COL["roas_7d"]),
        })

    return contents


# ─── KMS_CAL_DATA (Calendar view) ───────────────────────────

def parse_kms_calendar(rows, tab_label):
    """
    อ่านทุก row (ไม่ filter เฉพาะ TikTok) → สร้าง KMS_CAL_DATA entry
    รองรับทั้ง Published และ รอ Publish (แสดงใน calendar ได้ทั้งคู่)
    """
    year, month_idx = CAL_MONTH_MAP.get(tab_label, (__import__("datetime").datetime.now().year, __import__("datetime").datetime.now().month - 1))
    items = []

    for row in rows[1:]:  # ข้าม header
        creator = safe_get(row, COL["creator"])
        if creator not in CREATOR_NAMES:
            continue

        pub_status = safe_get(row, COL["published"])
        is_pub = pub_status in ("Published", "รอ Publish", "✓", "✅")

        # หา วัน/เดือน จาก pub_date ก่อน ถ้าไม่ได้ใช้ date
        pub_date_raw = safe_get(row, COL["pub_date"])
        date_raw     = safe_get(row, COL["date"])
        date_str     = parse_date(pub_date_raw, year) or parse_date(date_raw, year)
        if not date_str:
            continue
        try:
            day = int(date_str.split("-")[2])
        except:
            continue

        # TikTok URL (column 19 — full URL)
        tt_url = safe_get(row, COL["tiktok_url"])
        if tt_url and not tt_url.startswith("http"):
            tt_url = ""

        # Facebook URL: ใช้ fb_post_url โดยตรง หรือ สร้างจาก fb_post_id
        fb_url = safe_get(row, COL["fb_post_url"])
        fb_id  = safe_get(row, COL["fb_post_id"])
        if not fb_url and fb_id and re.match(r'^\d+$', fb_id):
            fb_url = f"https://www.facebook.com/opendurianhowto/posts/{fb_id}"

        # ROAS
        roas_raw = safe_get(row, COL["roas_7d"])
        try:
            roas = float(roas_raw)
        except:
            roas = None

        fmt       = safe_get(row, COL["format"]) or "VDO"
        threshold = 1.8 if fmt == "IMG" else 2.0
        success   = roas is not None and roas >= threshold

        items.append({
            "day":      day,
            "creator":  creator,
            "type":     fmt,
            "product":  safe_get(row, COL["product"]),
            "key_msg":  safe_get(row, COL["key_message"]),
            "pub":      is_pub,
            "pub_date": pub_date_raw,
            "new_msg":  False,
            "success":  success,
            "fb":       fb_url,
            "tt":       tt_url,
            "ads":      safe_get(row, COL["ads_fb_name"]),
            "roas":     roas,
        })

    label = f"{MONTH_LABELS.get(month_idx + 1, '')} {year}"
    print(f"  Calendar items = {len(items)} (เดือน {label})")
    return {
        "year":  year,
        "month": month_idx,   # 0-indexed (JS Date ใช้)
        "label": label,
        "items": items,
    }


# ─── helpers สำหรับ parse JS variable ────────────────────────

def extract_js_var(content, var_name):
    """
    ดึง JSON ของ var_name จาก JS source แบบนับ brace depth
    คืนค่า (json_str, start_idx, end_idx) หรือ (None, -1, -1) ถ้าไม่เจอ
    """
    marker = f"var {var_name} = "
    pos    = content.find(marker)
    if pos == -1:
        return None, -1, -1

    start   = pos + len(marker)
    open_ch = content[start] if start < len(content) else ""
    if open_ch not in ('{', '['):
        return None, -1, -1
    close_ch = '}' if open_ch == '{' else ']'

    depth, in_str, esc = 0, False, False
    for i, c in enumerate(content[start:]):
        if esc:
            esc = False; continue
        if c == '\\' and in_str:
            esc = True; continue
        if c == '"':
            in_str = not in_str; continue
        if not in_str:
            if c == open_ch:  depth += 1
            elif c == close_ch:
                depth -= 1
                if depth == 0:
                    end = start + i + 1
                    return content[start:end], start, end

    return None, -1, -1


# ─── updaters ────────────────────────────────────────────────

def update_dashboard_tiktok(new_tiktok_rows, html_path):
    """แทนที่แถว TikTok ใน RAW_DATA ด้วยข้อมูลใหม่"""
    if not os.path.exists(html_path):
        print(f"⚠️  ไม่พบ dashboard: {html_path}"); return

    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()

    # RAW_DATA ใช้ const (array)
    marker = "const RAW_DATA = "
    pos    = content.find(marker)
    if pos == -1:
        print("⚠️  ไม่พบ RAW_DATA"); return

    start = pos + len(marker)
    js_arr, _, end = extract_js_var(content, "RAW_DATA".replace("var ", ""))
    # fallback: ใช้ regex เดิมสำหรับ const
    match = re.search(r'const RAW_DATA = (\[.*?\]);', content, re.DOTALL)
    if not match:
        print("⚠️  parse RAW_DATA ไม่สำเร็จ"); return

    existing   = json.loads(match.group(1))
    meta_rows  = [r for r in existing if r.get("platform") != "TikTok"]
    merged     = meta_rows + new_tiktok_rows
    new_js     = json.dumps(merged, ensure_ascii=False, separators=(",", ":"))
    new_content = content[:match.start(1)] + new_js + content[match.end(1):]

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(new_content)

    print(f"✅ RAW_DATA อัปเดตแล้ว! TikTok={len(new_tiktok_rows)} Meta={len(meta_rows)}")


def update_kms_cal_data(new_cal, html_path):
    """
    อัปเดต KMS_CAL_DATA ใน index.html
    new_cal: dict { "Apr 26": {...}, ... }
    """
    if not os.path.exists(html_path):
        print(f"⚠️  ไม่พบ dashboard: {html_path}"); return

    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()

    js_str, s_idx, e_idx = extract_js_var(content, "KMS_CAL_DATA")
    if js_str is None:
        print("⚠️  ไม่พบ KMS_CAL_DATA ใน dashboard"); return

    try:
        existing = json.loads(js_str)
    except Exception as e:
        print(f"⚠️  parse KMS_CAL_DATA ล้มเหลว: {e}"); return

    # merge: เพิ่ม/อัปเดต months ที่มีข้อมูลใหม่
    for key, data in new_cal.items():
        if data.get("items"):
            existing[key] = data
            print(f"  อัปเดต KMS_CAL_DATA['{key}']: {len(data['items'])} items")

    # เรียงลำดับ chronological
    order = {"Jan 26":0,"Feb 26":1,"Mar 26":2,"Apr 26":3,"May 26":4,"Jun 26":5}
    sorted_cal = dict(sorted(existing.items(), key=lambda x: order.get(x[0], 9)))

    new_js      = json.dumps(sorted_cal, ensure_ascii=False, separators=(",", ":"))
    new_content = content[:s_idx] + new_js + content[e_idx:]

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(new_content)

    print(f"✅ KMS_CAL_DATA อัปเดตแล้ว! ({', '.join(sorted_cal.keys())})")


# ─── Supermetrics ────────────────────────────────────────────

def load_supermetrics(url):
    if not url:
        return {}
    rows = download_csv(url)
    if not rows:
        return {}

    header  = [h.strip().lower() for h in rows[0]]
    metrics = {}

    def col(name):
        for h in [name, name.replace("_",""), name.replace(" ","_").lower()]:
            if h in header:
                return header.index(h)
        return -1

    vid_col = col("video_id")
    if vid_col == -1:
        print("  ⚠️  ไม่พบคอลัมน์ video_id ใน Supermetrics Sheet")
        return {}

    for row in rows[1:]:
        if len(row) <= vid_col:
            continue
        vid = row[vid_col].strip()
        if not vid:
            continue
        def flt(c):
            try: return float(row[c]) if 0 <= c < len(row) else 0.0
            except: return 0.0
        if vid not in metrics:
            metrics[vid] = {"spend":0,"revenue":0,"purchases":0,
                            "impressions":0,"views_2s":0,"views_6s":0}
        m = metrics[vid]
        m["spend"]       += flt(col("spend"))
        m["revenue"]     += flt(col("revenue"))
        m["purchases"]   += flt(col("purchases"))
        m["impressions"] += flt(col("impressions"))
        m["views_2s"]    += flt(col("views_2s"))
        m["views_6s"]    += flt(col("views_6s"))

    return metrics


def build_raw_rows(contents, sm_metrics):
    rows = []
    for c in contents:
        vid     = c["video_id"]
        m       = sm_metrics.get(vid, {})
        spend   = m.get("spend", 0)
        revenue = m.get("revenue", 0)
        pur     = int(m.get("purchases", 0))
        imp     = int(m.get("impressions", 0))
        v2s     = m.get("views_2s", 0)
        v6s     = m.get("views_6s", 0)
        roas    = revenue / spend if spend > 0 else 0
        cpa     = spend / pur    if pur   > 0 else 0

        if roas == 0 and c["roas_7d"]:
            try: roas = float(c["roas_7d"])
            except: pass

        hit = "ไม่มียอดขาย"
        if pur > 0:
            threshold = 2.0 if c["content_format"] == "VDO" else 1.8
            hit = "ผ่าน" if roas >= threshold else "ไม่ผ่าน"

        rows.append({
            "month":          c["month"],
            "creator":        c["creator"],
            "ad_date":        c["date"],
            "content_name":   c["content_name"],
            "content_format": c["content_format"],
            "product":        c["product"],
            "platform":       "TikTok",
            "reach":          0,
            "impressions":    imp,
            "spend_thb":      round(spend, 2),
            "purchases":      pur,
            "revenue_thb":    round(revenue, 2),
            "roas":           round(roas, 4),
            "cpm_thb":        round(spend/imp*1000, 2) if imp > 0 else 0,
            "cpa_thb":        round(cpa, 2),
            "clicks":         0,
            "ctr":            0,
            "messages":       0,
            "v2s":            round(v2s, 2),
            "v6s":            round(v6s, 2),
            "ad_variants":    1,
            "hit_roas":       hit,
            "is_new":         False,
            "launch_date":    c["date"],
            "tiktok_video_id": vid,
        })
    return rows


# ─── main ────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  KMS Sheet Puller — OpenDurian How-to")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    all_contents = []
    cal_data     = {}           # { "Apr 26": {...} }
    downloaded   = {}          # cache rows เพื่อไม่ต้อง download ซ้ำ

    # auto-discover ทุกแท็บเดือนจาก workbook (xlsx) — ไม่ต้องใส่ GID มือ
    xlsx_tabs = discover_tabs_xlsx()
    tab_labels = sorted(set(list(xlsx_tabs.keys()) + [k for k, v in KMS_TAB_GIDS.items() if v]))
    print(f"\nAuto-discover แท็บ: {', '.join(tab_labels) if tab_labels else '(ไม่พบ)'}")

    for tab_label in tab_labels:
        gid = KMS_TAB_GIDS.get(tab_label, "")
        if gid:
            print(f"\n[{tab_label}] ดาวน์โหลด (GID/export csv)...")
            rows = download_csv(SHEET_CSV_URL.format(sid=KMS_SHEET_ID, gid=gid))
        else:
            print(f"\n[{tab_label}] อ่านจาก workbook (xlsx)...")
            rows = xlsx_tabs.get(tab_label, [])
        if not rows or len(rows) < 2:
            print(f"  ข้าม (ไม่มีข้อมูล/ยังไม่สร้างแท็บ)")
            continue
        downloaded[tab_label] = rows

        # --- RAW_DATA (TikTok content with video ID) ---
        contents = parse_kms_tab(rows, tab_label)
        print(f"  พบ {len(contents)} videos ที่มี TikTok ID")
        all_contents.extend(contents)

        # --- KMS_CAL_DATA (Calendar — ทุก row ที่มี creator + date) ---
        cal_key = CAL_KEY_MAP.get(tab_label, tab_label)
        cal_data[cal_key] = parse_kms_calendar(rows, tab_label)

    # ── อัปเดต RAW_DATA ──────────────────────────────────────
    if all_contents:
        print(f"\nรวม TikTok: {len(all_contents)} videos")
        by_creator = defaultdict(int)
        for c in all_contents:
            by_creator[c["creator"]] += 1
        for cr, n in sorted(by_creator.items()):
            print(f"  {cr}: {n} videos")

        map_path = os.path.join(os.path.dirname(DASHBOARD_PATH), "kms_content_map.json")
        with open(map_path, "w", encoding="utf-8") as f:
            json.dump(all_contents, f, ensure_ascii=False, indent=2)
        print(f"\n💾 Content map บันทึกแล้ว: kms_content_map.json")

        sm_metrics  = {}
        if SUPERMETRICS_CSV_URL:
            print("\nกำลังโหลด Supermetrics data...")
            sm_metrics = load_supermetrics(SUPERMETRICS_CSV_URL)
            print(f"  พบ metrics สำหรับ {len(sm_metrics)} videos")
        else:
            print("\n⚠️  ยังไม่มี Supermetrics URL — ใช้ ROAS จาก KMS แทน")

        tiktok_rows = build_raw_rows(all_contents, sm_metrics)
        update_dashboard_tiktok(tiktok_rows, DASHBOARD_PATH)
    else:
        print("\nไม่มีข้อมูล TikTok ใหม่ — ข้าม RAW_DATA update")

    # ── อัปเดต KMS_CAL_DATA (Calendar) ──────────────────────
    if cal_data:
        print(f"\n--- อัปเดต Calendar ---")
        update_kms_cal_data(cal_data, DASHBOARD_PATH)
    else:
        print("\nไม่มีข้อมูล Calendar ใหม่")


if __name__ == "__main__":
    main()
